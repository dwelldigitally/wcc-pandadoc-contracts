"""
Build WCC_Contract_Data.xlsx with the v3 column-based schema:
  1. Instructions (green tab)
  2. Programs (identity only)
  3. Intakes (core entity with hours, weeks, delivery)
  4. Fees (column-per-fee-type, one row per program+intake+residency)
  5. Outline Map

Run: python scripts/build-excel.py
Requires: pip install openpyxl
"""
from __future__ import annotations

import csv
import random
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

DATA_DIR = Path(__file__).resolve().parent.parent / "data"

# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
SCHOLARSHIP_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

# ---------------------------------------------------------------------------
# Dropdown options
# ---------------------------------------------------------------------------
CREDENTIALS = ["Diploma", "Certificate", "Bachelor", "Post-Graduate Diploma"]
CAMPUSES = ["Burnaby", "New Westminster", "Surrey", "Online"]
STATUSES = ["Open", "Closed", "Waitlist"]
DELIVERY_METHODS = ["In-Class", "Distance", "Combined", "Hybrid"]
RESIDENCY_OPTIONS = ["domestic", "international"]
SCHEDULES = [
    "Full-Time",
    "Part-Time",
    "Full-Time (Morning)",
    "Full-Time (Evening)",
    "Full-Time (Fast Track)",
]

# ---------------------------------------------------------------------------
# Canonical fee columns (left-to-right = display order on contract)
# ---------------------------------------------------------------------------
FEE_COLUMNS: list[str] = [
    "Application Fee",
    "Administration Fee",
    "Registration Fee",
    "Assessment Fee",
    "Tuition Fee",
    "Tuition Fee Per Credit",
    "Course Materials",
    "Textbooks",
    "Books",
    "Book Fee",
    "Annual Technology Fee",
    "Ground School Fee",
    "Flight Dual",
    "Flight Solo",
    "Flight Time Building",
    "Flight Prep Ground",
    "Fuel",
    "Other",
    "Scholarship",
]

# Map raw CSV fee_name → canonical column name
FEE_NAME_MAP: dict[str, str] = {
    "Application Fee": "Application Fee",
    "Administration Fee": "Administration Fee",
    "Administrative Fee": "Administration Fee",
    "Registration Fee": "Registration Fee",
    "Assessment Fee": "Assessment Fee",
    "Tuition Fee": "Tuition Fee",
    "Tuition Fee Per Credit": "Tuition Fee Per Credit",
    "Course Materials": "Course Materials",
    "Course Materials Fee": "Course Materials",
    "Books": "Books",
    "Book Fee": "Book Fee",
    "Textbooks": "Textbooks",
    "Textbooks Fee": "Textbooks",
    "Text Books": "Textbooks",
    "Text Books Fee": "Textbooks",
    "Ground School Fee": "Ground School Fee",
    "Dual (30 hrs) @$290/hr": "Flight Dual",
    "Dual (35 hrs) @$290/hr": "Flight Dual",
    "Solo (15 hrs) @$225/hr": "Flight Solo",
    "Solo (30 hrs) @$225/hr": "Flight Solo",
    "Time Building (90 hrs) @$225/hr": "Flight Time Building",
    "Preparatory Ground (PGI/GB)": "Flight Prep Ground",
    "Fuel ($50 x 45 hours)": "Fuel",
    "Fuel ($50 x 155 hours)": "Fuel",
    "Annual Technology Fee": "Annual Technology Fee",
    "Other": "Other",
    "Other Fee": "Other",
    "Other Materials": "Other",
    "Scholarship": "Scholarship",
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def style_header(ws: Worksheet, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def auto_width(ws: Worksheet) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def load_csv(filepath: Path) -> tuple[list[str], list[dict[str, str]]]:
    with open(filepath, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return list(reader.fieldnames or []), list(reader)


def col_range(col_letter: str, start: int = 2, end: int = 500) -> str:
    return f"{col_letter}{start}:{col_letter}{end}"


def add_dropdown(
    ws: Worksheet,
    cell_range: str,
    options: list[str],
    *,
    allow_blank: bool = True,
    error_title: str = "Invalid Entry",
    error_msg: str = "Please select a value from the dropdown list.",
) -> None:
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=allow_blank,
        showErrorMessage=True,
        errorTitle=error_title,
        error=error_msg,
        showInputMessage=True,
        promptTitle="Select",
        prompt="Pick from dropdown",
    )
    dv.sqref = cell_range
    ws.add_data_validation(dv)


def add_ref_dropdown(
    ws: Worksheet,
    cell_range: str,
    ref_formula: str,
    *,
    allow_blank: bool = False,
    error_title: str = "Invalid Program Name",
    error_msg: str = "Value must match a program_name in the Programs tab.",
) -> None:
    dv = DataValidation(
        type="list",
        formula1=ref_formula,
        allow_blank=allow_blank,
        showErrorMessage=True,
        errorTitle=error_title,
        error=error_msg,
        showInputMessage=True,
        promptTitle="Program",
        prompt="Select a program name",
    )
    dv.sqref = cell_range
    ws.add_data_validation(dv)


def add_number_validation(
    ws: Worksheet,
    cell_range: str,
    *,
    min_val: float = 0,
    max_val: float = 999999,
    allow_blank: bool = True,
    is_whole: bool = False,
    error_title: str = "Invalid Number",
    error_msg: str | None = None,
) -> None:
    val_type = "whole" if is_whole else "decimal"
    msg = error_msg or f"Enter a number between {min_val} and {max_val}."
    dv = DataValidation(
        type=val_type,
        operator="between",
        formula1=str(min_val),
        formula2=str(max_val),
        allow_blank=allow_blank,
        showErrorMessage=True,
        errorTitle=error_title,
        error=msg,
    )
    dv.sqref = cell_range
    ws.add_data_validation(dv)


def add_date_validation(
    ws: Worksheet,
    cell_range: str,
    *,
    allow_blank: bool = False,
) -> None:
    dv = DataValidation(
        type="date",
        operator="between",
        formula1=datetime(2020, 1, 1),
        formula2=datetime(2099, 12, 31),
        allow_blank=allow_blank,
        showErrorMessage=True,
        errorTitle="Invalid Date",
        error="Enter a valid date (YYYY-MM-DD format).",
        showInputMessage=True,
        promptTitle="Date",
        prompt="Enter date as YYYY-MM-DD",
    )
    dv.sqref = cell_range
    ws.add_data_validation(dv)


# ---------------------------------------------------------------------------
# Pivot fees from row-based CSV into column-based dicts
# ---------------------------------------------------------------------------


def pivot_fees(
    fees_raw: list[dict[str, str]],
) -> dict[str, dict[str, float]]:
    """
    Returns { program_name: { "domestic": {col: amount}, "international": {col: amount} } }
    Sums amounts when multiple CSV rows map to the same canonical column.
    """
    result: dict[str, dict[str, dict[str, float]]] = {}

    for fee in fees_raw:
        prog = fee["program_name"]
        raw_name = fee["fee_name"]
        canonical = FEE_NAME_MAP.get(raw_name, raw_name)

        if canonical not in FEE_COLUMNS:
            print(f"  WARNING: unmapped fee '{raw_name}' for {prog} → using 'Other'")
            canonical = "Other"

        dom_amt = _parse_amount(fee.get("domestic_amount", ""))
        intl_amt = _parse_amount(fee.get("international_amount", ""))

        if prog not in result:
            result[prog] = {"domestic": {}, "international": {}}

        result[prog]["domestic"][canonical] = (
            result[prog]["domestic"].get(canonical, 0) + dom_amt
        )
        result[prog]["international"][canonical] = (
            result[prog]["international"].get(canonical, 0) + intl_amt
        )

    return result


def _parse_amount(val: str) -> float:
    cleaned = val.replace("$", "").replace(",", "").strip()
    if not cleaned:
        return 0
    try:
        return float(cleaned)
    except ValueError:
        return 0


# ---------------------------------------------------------------------------
# Generate dummy intakes
# ---------------------------------------------------------------------------

# Programs CSV still has hours/weeks — pull them for intake generation
PROGRAM_HOURS: dict[str, dict[str, Any]] = {}


def generate_intakes(programs_raw: list[dict[str, str]]) -> list[dict[str, Any]]:
    random.seed(42)
    intakes: list[dict[str, Any]] = []

    for prog in programs_raw:
        name = prog["program_name"]

        # Parse hours/weeks from old CSV for seeding
        hours = _safe_int(prog.get("hours", ""))
        weeks = _safe_int(prog.get("weeks_full_time", ""))
        PROGRAM_HOURS[name] = {"hours": hours, "weeks": weeks}

        old_delivery = prog.get("delivery_method", "In-Class")
        dom_delivery = "In-Class"
        if old_delivery and "distance" in old_delivery.lower():
            dom_delivery = "Distance"
        intl_delivery = "In-Class"

        duration_weeks = weeks if weeks and weeks > 0 else 24

        num_intakes = random.choice([2, 2, 3])
        possible_starts = [
            datetime(2026, 4, random.choice([7, 14, 21])),
            datetime(2026, 5, random.choice([5, 12, 19])),
            datetime(2026, 9, random.choice([8, 15, 22])),
            datetime(2026, 10, random.choice([5, 12])),
            datetime(2027, 1, random.choice([6, 13, 20])),
            datetime(2027, 2, random.choice([3, 10, 17])),
        ]
        chosen = sorted(
            random.sample(possible_starts, min(num_intakes, len(possible_starts)))
        )

        for start in chosen:
            end = start + timedelta(weeks=duration_weeks)
            campus = random.choice(CAMPUSES)
            status = random.choice(["Open", "Open", "Open", "Closed", "Waitlist"])
            spots = (
                random.randint(2, 25)
                if status == "Open"
                else (0 if status == "Closed" else random.randint(0, 3))
            )
            intakes.append(
                {
                    "program_name": name,
                    "intake_date": start.strftime("%Y-%m-%d"),
                    "end_date": end.strftime("%Y-%m-%d"),
                    "campus": campus,
                    "schedule": random.choice(["Full-Time", "Part-Time"]),
                    "hours": hours or "",
                    "weeks": weeks or "",
                    "domestic_delivery_method": dom_delivery,
                    "international_delivery_method": intl_delivery,
                    "spots_available": spots,
                    "status": status,
                }
            )

    intakes.sort(key=lambda x: (x["program_name"], x["intake_date"]))
    return intakes


def _safe_int(val: str) -> int | None:
    try:
        return int(float(val)) if val else None
    except (ValueError, TypeError):
        return None


# ===========================================================================
# Main
# ===========================================================================


def main() -> None:
    # Load source CSV data
    _, programs_raw = load_csv(DATA_DIR / "demo-programs.csv")
    _, fees_raw = load_csv(DATA_DIR / "demo-fees.csv")
    _, outlines_raw = load_csv(DATA_DIR / "program-outline-map.csv")

    intakes = generate_intakes(programs_raw)
    pivoted = pivot_fees(fees_raw)
    print(f"Loaded {len(programs_raw)} programs, {len(fees_raw)} fee rows, {len(intakes)} intakes")

    wb = openpyxl.Workbook()

    # ==================================================================
    # TAB 1 — Instructions
    # ==================================================================
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    ws_instr.sheet_properties.tabColor = "00B050"

    instructions = [
        ("WCC Contract Data — Instructions", Font(bold=True, size=14)),
        ("", None),
        ("This workbook feeds the automated contract generation system.", None),
        ("Changes here take effect IMMEDIATELY for the next contract generated.", None),
        ("", None),
        ("PROGRAMS TAB", Font(bold=True, size=11, color="1F4E79")),
        ("- program_name: Must EXACTLY match the HubSpot 'Program of Study' property", None),
        ("- credential: Use the dropdown (Diploma, Certificate, Bachelor, Post-Graduate Diploma)", None),
        ("- active: Set to FALSE to disable a program", None),
        ("- Note: hours, weeks, delivery method are on the Intakes tab (per cohort)", None),
        ("", None),
        ("FEES TAB (column-based layout)", Font(bold=True, size=11, color="1F4E79")),
        ("- One row per program + effective_from date + residency tier (domestic / international)", None),
        ("- Each fee type is a COLUMN. Leave blank if N/A for that program.", None),
        ("- Amounts are numbers only — NO $ signs, NO commas", None),
        ("- Column order = display order on the contract (left to right)", None),
        ("- Scholarship column: ALWAYS a negative number (reduces the total)", None),
        ("- Total column: Excel SUM formula — do NOT edit, it auto-calculates", None),
        ("- To add a NEW fee type: insert a column before Scholarship", None),
        ("- To update fees: add new rows with a later effective_from date (old rows still apply to earlier intakes)", None),
        ("", None),
        ("INTAKES TAB", Font(bold=True, size=11, color="1F4E79")),
        ("- One row per intake (enrollment window) per program", None),
        ("- hours, weeks, delivery method are PER COHORT — not per program", None),
        ("- domestic_delivery_method / international_delivery_method can differ for the same intake", None),
        ("- intake_date / end_date format: YYYY-MM-DD", None),
        ("- status: Open = available, Closed = hidden from advisors, Waitlist = full but accepting", None),
        ("", None),
        ("CONDITIONAL FORMATTING LEGEND", Font(bold=True, size=11, color="1F4E79")),
        ("  RED cells = blank required field or Closed status", None),
        ("  YELLOW cells = $0 amount (verify if intentional) or Waitlist status", None),
        ("  GREEN cells = Open status", None),
        ("  BLUE cells = Scholarship (negative amount)", None),
        ("", None),
        ("COMMON MISTAKES TO AVOID", Font(bold=True, size=11, color="CC0000")),
        ("- Typo in program_name → fees not found, advisor sees $0 total", None),
        ("- Missing intake row → no fees for that cohort", None),
        ("- Positive number in Scholarship → validation will reject it", None),
        ("- Wrong date format → use YYYY-MM-DD not MM/DD/YYYY", None),
        ("- Adding $ sign or comma in fee amounts → validation will reject it", None),
        ("- Forgetting both domestic AND international rows for an intake", None),
        ("", None),
        (f"Last updated: {datetime.now().strftime('%Y-%m-%d')} — v3 schema (column-based fees, scholarship support)", None),
    ]

    for ri, (text, font) in enumerate(instructions, 1):
        cell = ws_instr.cell(row=ri, column=1, value=text)
        if font:
            cell.font = font
    ws_instr.column_dimensions["A"].width = 85

    # ==================================================================
    # TAB 2 — Programs (identity only)
    # ==================================================================
    ws_prog = wb.create_sheet("Programs")
    prog_headers = ["program_name", "program_code", "credential", "active"]
    for ci, h in enumerate(prog_headers, 1):
        ws_prog.cell(row=1, column=ci, value=h)

    for ri, prog in enumerate(programs_raw, 2):
        ws_prog.cell(row=ri, column=1, value=prog.get("program_name", ""))
        ws_prog.cell(row=ri, column=2, value=prog.get("program_code", ""))
        ws_prog.cell(row=ri, column=3, value=prog.get("credential", ""))
        ws_prog.cell(row=ri, column=4, value="TRUE")

    style_header(ws_prog, len(prog_headers))
    auto_width(ws_prog)

    # Validations
    add_dropdown(
        ws_prog,
        col_range("C"),
        CREDENTIALS,
        allow_blank=False,
        error_title="Invalid Credential",
        error_msg="Must be: Diploma, Certificate, Bachelor, or Post-Graduate Diploma",
    )
    add_dropdown(
        ws_prog,
        col_range("D"),
        ["TRUE", "FALSE"],
        allow_blank=False,
        error_title="Invalid Active",
        error_msg="Must be TRUE or FALSE",
    )

    # Conditional formatting — blank program_name
    ws_prog.conditional_formatting.add(
        col_range("A"), FormulaRule(formula=["ISBLANK(A2)"], fill=RED_FILL)
    )

    # ==================================================================
    # TAB 3 — Intakes (core entity)
    # ==================================================================
    ws_intakes = wb.create_sheet("Intakes")
    intake_headers = [
        "program_name",       # A
        "intake_date",        # B
        "end_date",           # C
        "campus",             # D
        "schedule",           # E
        "hours",              # F
        "weeks",              # G
        "domestic_delivery_method",  # H
        "international_delivery_method",  # I
        "spots_available",    # J
        "status",             # K
    ]
    for ci, h in enumerate(intake_headers, 1):
        ws_intakes.cell(row=1, column=ci, value=h)

    for ri, intake in enumerate(intakes, 2):
        ws_intakes.cell(row=ri, column=1, value=intake["program_name"])
        ws_intakes.cell(row=ri, column=2, value=intake["intake_date"])
        ws_intakes.cell(row=ri, column=3, value=intake["end_date"])
        ws_intakes.cell(row=ri, column=4, value=intake["campus"])
        ws_intakes.cell(row=ri, column=5, value=intake["schedule"])
        h_val = intake["hours"]
        ws_intakes.cell(row=ri, column=6, value=h_val if h_val else None)
        w_val = intake["weeks"]
        ws_intakes.cell(row=ri, column=7, value=w_val if w_val else None)
        ws_intakes.cell(row=ri, column=8, value=intake["domestic_delivery_method"])
        ws_intakes.cell(row=ri, column=9, value=intake["international_delivery_method"])
        ws_intakes.cell(row=ri, column=10, value=int(intake["spots_available"]))
        ws_intakes.cell(row=ri, column=11, value=intake["status"])

    style_header(ws_intakes, len(intake_headers))
    auto_width(ws_intakes)

    MR = 500  # max validation row

    # --- Intakes validations ---
    add_ref_dropdown(ws_intakes, col_range("A", 2, MR), "=Programs!$A$2:$A$200")
    add_date_validation(ws_intakes, col_range("B", 2, MR))
    add_date_validation(ws_intakes, col_range("C", 2, MR))
    add_dropdown(
        ws_intakes, col_range("D", 2, MR), CAMPUSES,
        allow_blank=False, error_title="Invalid Campus", error_msg="Select a campus",
    )
    add_dropdown(
        ws_intakes, col_range("E", 2, MR), SCHEDULES,
        allow_blank=False, error_title="Invalid Schedule", error_msg="Select a schedule type",
    )
    add_number_validation(
        ws_intakes, col_range("F", 2, MR), min_val=1, max_val=99999,
        error_title="Invalid Hours", error_msg="Hours must be > 0",
    )
    add_number_validation(
        ws_intakes, col_range("G", 2, MR), min_val=1, max_val=9999,
        error_title="Invalid Weeks", error_msg="Weeks must be > 0",
    )
    add_dropdown(
        ws_intakes, col_range("H", 2, MR), DELIVERY_METHODS,
        error_title="Invalid Delivery", error_msg="Select a delivery method",
    )
    add_dropdown(
        ws_intakes, col_range("I", 2, MR), DELIVERY_METHODS,
        error_title="Invalid Delivery", error_msg="Select a delivery method",
    )
    add_number_validation(
        ws_intakes, col_range("J", 2, MR), min_val=0, max_val=999,
        is_whole=True, error_title="Invalid Spots", error_msg="Spots must be >= 0",
    )
    add_dropdown(
        ws_intakes, col_range("K", 2, MR), STATUSES,
        allow_blank=False, error_title="Invalid Status", error_msg="Must be Open, Closed, or Waitlist",
    )

    # --- Intakes conditional formatting ---
    ws_intakes.conditional_formatting.add(
        col_range("A", 2, MR), FormulaRule(formula=["ISBLANK(A2)"], fill=RED_FILL)
    )
    ws_intakes.conditional_formatting.add(
        col_range("B", 2, MR), FormulaRule(formula=["ISBLANK(B2)"], fill=RED_FILL)
    )
    ws_intakes.conditional_formatting.add(
        col_range("K", 2, MR),
        CellIsRule(operator="equal", formula=['"Open"'], fill=GREEN_FILL),
    )
    ws_intakes.conditional_formatting.add(
        col_range("K", 2, MR),
        CellIsRule(operator="equal", formula=['"Closed"'], fill=RED_FILL),
    )
    ws_intakes.conditional_formatting.add(
        col_range("K", 2, MR),
        CellIsRule(operator="equal", formula=['"Waitlist"'], fill=YELLOW_FILL),
    )

    # ==================================================================
    # TAB 4 — Fees (column-per-fee-type)
    # ==================================================================
    ws_fees = wb.create_sheet("Fees")

    # Headers: key columns + fee columns + Total
    fee_headers = ["program_name", "effective_from", "residency"] + FEE_COLUMNS + ["Total"]
    for ci, h in enumerate(fee_headers, 1):
        ws_fees.cell(row=1, column=ci, value=h)

    # Highlight Scholarship header in blue
    scholarship_col_idx = fee_headers.index("Scholarship") + 1
    ws_fees.cell(row=1, column=scholarship_col_idx).fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    # Highlight Total header differently
    total_col_idx = fee_headers.index("Total") + 1
    ws_fees.cell(row=1, column=total_col_idx).fill = PatternFill(
        start_color="548235", end_color="548235", fill_type="solid"
    )

    # Populate rows: one domestic + one international row per program
    # effective_from defaults to 2026-01-01 (from CSV data)
    # Finance adds new rows with later effective_from when fees change
    row_num = 2
    effective_from = "2026-01-01"

    for prog in programs_raw:
        prog_name = prog["program_name"]
        prog_fees = pivoted.get(prog_name, {"domestic": {}, "international": {}})

        for residency in ["domestic", "international"]:
            ws_fees.cell(row=row_num, column=1, value=prog_name)
            ws_fees.cell(row=row_num, column=2, value=effective_from)
            ws_fees.cell(row=row_num, column=3, value=residency)

            tier_fees = prog_fees.get(residency, {})
            for fi, fee_col in enumerate(FEE_COLUMNS):
                col_idx = 4 + fi  # offset by 3 key columns
                amount = tier_fees.get(fee_col, None)
                if amount and amount != 0:
                    cell = ws_fees.cell(row=row_num, column=col_idx, value=amount)
                    cell.number_format = '#,##0'

            # Total = SUM of all fee columns in this row
            first_fee_letter = get_column_letter(4)
            last_fee_letter = get_column_letter(3 + len(FEE_COLUMNS))
            total_col = 4 + len(FEE_COLUMNS)
            total_cell = ws_fees.cell(
                row=row_num,
                column=total_col,
                value=f"=SUM({first_fee_letter}{row_num}:{last_fee_letter}{row_num})",
            )
            total_cell.number_format = '#,##0'
            total_cell.font = Font(bold=True)

            row_num += 1

    style_header(ws_fees, len(fee_headers))
    auto_width(ws_fees)

    # Make key columns wider for readability
    ws_fees.column_dimensions["A"].width = 42  # program_name
    ws_fees.column_dimensions["B"].width = 14  # intake_date
    ws_fees.column_dimensions["C"].width = 14  # residency

    max_fee_row = max(row_num + 100, 500)

    # --- Fees validations ---

    # program_name dropdown from Programs tab
    add_ref_dropdown(ws_fees, col_range("A", 2, max_fee_row), "=Programs!$A$2:$A$200")

    # intake_date — date validation
    add_date_validation(ws_fees, col_range("B", 2, max_fee_row))

    # residency dropdown
    add_dropdown(
        ws_fees, col_range("C", 2, max_fee_row), RESIDENCY_OPTIONS,
        allow_blank=False,
        error_title="Invalid Residency",
        error_msg="Must be 'domestic' or 'international'",
    )

    # Fee amount columns: must be >= 0 (except Scholarship)
    for fi, fee_col in enumerate(FEE_COLUMNS):
        col_letter = get_column_letter(4 + fi)
        cell_rng = col_range(col_letter, 2, max_fee_row)

        if fee_col == "Scholarship":
            # Scholarship: must be <= 0
            add_number_validation(
                ws_fees, cell_rng,
                min_val=-999999, max_val=0,
                error_title="Invalid Scholarship",
                error_msg="Scholarship must be negative or zero (e.g., -2000)",
            )
        else:
            # Regular fee: must be >= 0
            add_number_validation(
                ws_fees, cell_rng,
                min_val=0, max_val=999999,
                error_title="Invalid Fee Amount",
                error_msg="Fee amount must be a number >= 0. No $ signs.",
            )

    # --- Fees conditional formatting ---

    # Blank required: program_name (A), intake_date (B), residency (C)
    for col_letter in ["A", "B", "C"]:
        ws_fees.conditional_formatting.add(
            col_range(col_letter, 2, max_fee_row),
            FormulaRule(formula=[f"ISBLANK({col_letter}2)"], fill=RED_FILL),
        )

    # Scholarship column: highlight blue when populated (non-zero)
    schol_letter = get_column_letter(scholarship_col_idx)
    ws_fees.conditional_formatting.add(
        col_range(schol_letter, 2, max_fee_row),
        FormulaRule(formula=[f"{schol_letter}2<0"], fill=SCHOLARSHIP_FILL),
    )

    # Highlight entire row blue when Scholarship is present
    all_fee_range = f"A2:{get_column_letter(total_col_idx)}{max_fee_row}"
    ws_fees.conditional_formatting.add(
        all_fee_range,
        FormulaRule(
            formula=[f"${schol_letter}2<0"],
            fill=PatternFill(start_color="EDF2F9", end_color="EDF2F9", fill_type="solid"),
        ),
    )

    # Total column: bold green when > 0
    total_letter = get_column_letter(total_col_idx)
    ws_fees.conditional_formatting.add(
        col_range(total_letter, 2, max_fee_row),
        FormulaRule(formula=[f"{total_letter}2>0"], fill=GREEN_FILL),
    )

    # ==================================================================
    # TAB 5 — Outline Map
    # ==================================================================
    ws_outline = wb.create_sheet("Outline Map")
    outline_headers = ["program_name", "outline_filename"]
    for ci, h in enumerate(outline_headers, 1):
        ws_outline.cell(row=1, column=ci, value=h)

    for ri, row in enumerate(outlines_raw, 2):
        ws_outline.cell(row=ri, column=1, value=row.get("program_name", ""))
        ws_outline.cell(row=ri, column=2, value=row.get("outline_filename", ""))

    style_header(ws_outline, len(outline_headers))
    auto_width(ws_outline)

    add_ref_dropdown(ws_outline, col_range("A", 2, 200), "=Programs!$A$2:$A$200")
    ws_outline.conditional_formatting.add(
        col_range("A", 2, 200),
        FormulaRule(formula=["ISBLANK(A2)"], fill=RED_FILL),
    )

    # ==================================================================
    # Save
    # ==================================================================
    outpath = DATA_DIR / "WCC_Contract_Data.xlsx"
    wb.save(outpath)

    # Report
    prog_count = ws_prog.max_row - 1
    intake_count = ws_intakes.max_row - 1
    fee_count = ws_fees.max_row - 1
    outline_count = ws_outline.max_row - 1

    print(f"\nSaved: {outpath}")
    print(f"  Instructions:  info tab (green)")
    print(f"  Programs:      {prog_count} rows")
    print(f"  Intakes:       {intake_count} rows")
    print(f"  Fees:          {fee_count} rows ({fee_count // 2} programs x 2 residency tiers)")
    print(f"  Outline Map:   {outline_count} rows")
    print(f"  Fee columns:   {len(FEE_COLUMNS)} types + Total")
    print(f"\nValidation rules applied:")
    print(f"  Programs:      {len(ws_prog.data_validations.dataValidation)}")
    print(f"  Intakes:       {len(ws_intakes.data_validations.dataValidation)}")
    print(f"  Fees:          {len(ws_fees.data_validations.dataValidation)}")
    print(f"  Outline Map:   {len(ws_outline.data_validations.dataValidation)}")


if __name__ == "__main__":
    main()
