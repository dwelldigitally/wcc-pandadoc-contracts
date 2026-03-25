"""Add data validation rules and instructions to the WCC Contract Data Excel file."""

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule
from pathlib import Path

DATA_DIR = Path(__file__).parent.parent / "data"
wb = openpyxl.load_workbook(DATA_DIR / "WCC_Contract_Data.xlsx")

red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

# ==========================================
# PROGRAMS TAB
# ==========================================
ws = wb["Programs"]

# Credential dropdown (column D)
dv = DataValidation(type="list", formula1='"Diploma,Certificate,Bachelor,Post-Graduate Diploma"',
    allow_blank=False, showErrorMessage=True,
    errorTitle="Invalid Credential", error="Must be: Diploma, Certificate, Bachelor, or Post-Graduate Diploma")
ws.add_data_validation(dv)
dv.add(f"D2:D{ws.max_row + 50}")

# Delivery method dropdown (column I)
dv = DataValidation(type="list", formula1='"In-class,Distance-Synchronous,Distance-Asynchronous,Combined"',
    allow_blank=True, showErrorMessage=True,
    errorTitle="Invalid Delivery Method", error="Must be: In-class, Distance-Synchronous, Distance-Asynchronous, or Combined")
ws.add_data_validation(dv)
dv.add(f"I2:I{ws.max_row + 50}")

# Active dropdown (column J)
dv = DataValidation(type="list", formula1='"TRUE,FALSE"',
    allow_blank=False, showErrorMessage=True,
    errorTitle="Invalid Active", error="Must be TRUE or FALSE")
ws.add_data_validation(dv)
dv.add(f"J2:J{ws.max_row + 50}")

# Hours must be numeric (column E)
dv = DataValidation(type="decimal", operator="greaterThan", formula1="0",
    allow_blank=True, showErrorMessage=True,
    errorTitle="Invalid Hours", error="Hours must be a number greater than 0")
ws.add_data_validation(dv)
dv.add(f"E2:E{ws.max_row + 50}")

# Weeks must be numeric (columns F, G)
dv = DataValidation(type="decimal", operator="greaterThan", formula1="0",
    allow_blank=True, showErrorMessage=True,
    errorTitle="Invalid Weeks", error="Weeks must be a number greater than 0")
ws.add_data_validation(dv)
dv.add(f"F2:F{ws.max_row + 50}")
dv.add(f"G2:G{ws.max_row + 50}")

# Red highlight blank program names
ws.conditional_formatting.add(f"A2:A{ws.max_row + 50}",
    CellIsRule(operator="equal", formula=['""'], fill=red_fill))

# ==========================================
# FEES TAB
# ==========================================
ws_fees = wb["Fees"]

# Fee amounts must be >= 0 (columns C, D)
dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0",
    allow_blank=False, showErrorMessage=True,
    errorTitle="Invalid Fee Amount", error="Fee amount must be a number >= 0. No $ signs.")
ws_fees.add_data_validation(dv)
dv.add(f"C2:C{ws_fees.max_row + 100}")
dv.add(f"D2:D{ws_fees.max_row + 100}")

# is_tuition dropdown (column E)
dv = DataValidation(type="list", formula1='"TRUE,FALSE"',
    allow_blank=False, showErrorMessage=True,
    errorTitle="Invalid is_tuition", error="Must be TRUE or FALSE")
ws_fees.add_data_validation(dv)
dv.add(f"E2:E{ws_fees.max_row + 100}")

# sort_order must be integer >= 1 (column F)
dv = DataValidation(type="whole", operator="greaterThan", formula1="0",
    allow_blank=False, showErrorMessage=True,
    errorTitle="Invalid Sort Order", error="Sort order must be a whole number >= 1")
ws_fees.add_data_validation(dv)
dv.add(f"F2:F{ws_fees.max_row + 100}")

# Red highlight blank required fields
ws_fees.conditional_formatting.add(f"A2:A{ws_fees.max_row + 100}",
    CellIsRule(operator="equal", formula=['""'], fill=red_fill))
ws_fees.conditional_formatting.add(f"B2:B{ws_fees.max_row + 100}",
    CellIsRule(operator="equal", formula=['""'], fill=red_fill))

# Yellow highlight $0 amounts (review if intentional)
ws_fees.conditional_formatting.add(f"C2:C{ws_fees.max_row + 100}",
    CellIsRule(operator="equal", formula=["0"], fill=yellow_fill))
ws_fees.conditional_formatting.add(f"D2:D{ws_fees.max_row + 100}",
    CellIsRule(operator="equal", formula=["0"], fill=yellow_fill))

# ==========================================
# INTAKES TAB
# ==========================================
ws_int = wb["Intakes"]

# Status dropdown (column G)
dv = DataValidation(type="list", formula1='"Open,Closed,Waitlist"',
    allow_blank=False, showErrorMessage=True,
    errorTitle="Invalid Status", error="Must be: Open, Closed, or Waitlist")
ws_int.add_data_validation(dv)
dv.add(f"G2:G{ws_int.max_row + 200}")

# Schedule dropdown (column E)
dv = DataValidation(type="list",
    formula1='"Full-Time,Part-Time,Full-Time (Morning),Full-Time (Evening),Full-Time (Fast Track)"',
    allow_blank=False, showErrorMessage=True,
    errorTitle="Invalid Schedule", error="Must be: Full-Time, Part-Time, etc.")
ws_int.add_data_validation(dv)
dv.add(f"E2:E{ws_int.max_row + 200}")

# Spots must be >= 0 (column F)
dv = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0",
    allow_blank=True, showErrorMessage=True,
    errorTitle="Invalid Spots", error="Spots must be a whole number >= 0")
ws_int.add_data_validation(dv)
dv.add(f"F2:F{ws_int.max_row + 200}")

# Red highlight blank program names
ws_int.conditional_formatting.add(f"A2:A{ws_int.max_row + 200}",
    CellIsRule(operator="equal", formula=['""'], fill=red_fill))

# ==========================================
# INSTRUCTIONS TAB
# ==========================================
ws_help = wb.create_sheet("Instructions", 0)

instructions = [
    "WCC Contract Data - Instructions",
    "",
    "This spreadsheet feeds the automated contract generation system.",
    "Changes here take effect IMMEDIATELY for the next contract generated.",
    "",
    "PROGRAMS TAB",
    "- program_name: Must EXACTLY match the HubSpot 'Program of Study' property",
    "- hours, weeks: Numbers only, no text",
    "- credential: Use the dropdown (Diploma, Certificate, Bachelor, Post-Graduate Diploma)",
    "- active: Set to FALSE to disable a program",
    "",
    "FEES TAB",
    "- program_name: Must EXACTLY match the Programs tab (copy-paste recommended)",
    "- domestic_amount / international_amount: Numbers only, NO $ signs",
    "- Yellow highlight = $0 amount (review if intentional)",
    "- Red highlight = blank required field",
    "- sort_order: Controls display order on the contract (1 = first)",
    "",
    "INTAKES TAB",
    "- program_name: Must EXACTLY match the Programs tab",
    "- intake_date / end_date: Use YYYY-MM-DD format (e.g., 2026-05-05)",
    "- status: Open = available for new contracts, Closed = hidden from advisors",
    "- Only OPEN intakes with future dates appear in the advisor's picker",
    "",
    "OUTLINE MAP TAB",
    "- Maps programs to their PDF outline files",
    "- outline_filename: The filename WITHOUT the .pdf extension",
    "- Do not edit unless new program outlines are added",
    "",
    "COMMON MISTAKES TO AVOID",
    "- Typo in program_name -> fees not found, advisor sees $0 total",
    "- Missing fee row -> program shows fewer items than expected",
    "- Leaving amount blank -> contract shows $0",
    "- Wrong date format -> use YYYY-MM-DD not MM/DD/YYYY",
    "- Forgetting to set status to Open on new intakes",
    "- Adding $ sign in fee amounts -> validation will reject it",
]

bold_font = Font(bold=True, size=14)
header_font = Font(bold=True, size=11, color="1F4E79")

for i, text in enumerate(instructions):
    ws_help.cell(row=i + 1, column=1, value=text)
    if i == 0:
        ws_help.cell(row=i + 1, column=1).font = bold_font
    elif text.endswith("TAB") or text == "COMMON MISTAKES TO AVOID":
        ws_help.cell(row=i + 1, column=1).font = header_font

ws_help.column_dimensions["A"].width = 80

# Save
wb.save(DATA_DIR / "WCC_Contract_Data.xlsx")
print("Done! Added:")
print("  - Instructions tab (first tab)")
print("  - Dropdown validations: credential, delivery, active, status, schedule, is_tuition")
print("  - Numeric validations: hours, weeks, fee amounts, sort_order, spots")
print("  - Red highlight on blank required fields")
print("  - Yellow highlight on $0 fee amounts")
